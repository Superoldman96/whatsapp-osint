from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sqlite3
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

class Converter:
    def __init__(self, db_path: str = 'database/victims_logs.db', excel_file: str = 'History_wp.xlsx'):
        self.db_path = Path(db_path)
        self.excel_file = Path(excel_file)

    def db_to_excel(self):
        """Exports data from the database to an Excel file."""
        if not self.db_path.exists():
            logger.error(f"Database not found at {self.db_path}")
            return

        conn = None
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            wb = Workbook()
            ws = wb.active

            # Styles
            bold = Font(bold=True, name='Arial', color="00800000", size=10)
            align = Alignment(horizontal="center")

            # Column Configuration
            headers = [
                ("A", 15, "Session ID"),
                ("B", 17, "Username"),
                ("C", 20, "Start DateTime"),
                ("D", 20, "End DateTime"),
                ("E", 15, "Time Connected (s)")
            ]

            for col, width, title in headers:
                ws.column_dimensions[col].width = width
                cell = ws[f"{col}1"]
                cell.font = bold
                cell.alignment = align
                cell.value = title

            ws.title = "History Of Their Wp"

            # Query to join Users and Sessions
            query = '''
                SELECT
                    s.id,
                    u.user_name,
                    s.start_date || ' ' || s.start_hour || ':' || s.start_minute || ':' || s.start_second AS start_datetime,
                    s.end_date || ' ' || s.end_hour || ':' || s.end_minute || ':' || s.end_second AS end_datetime,
                    s.time_connected
                FROM Sessions s
                JOIN Users u ON s.user_id = u.id
                WHERE s.end_date IS NOT NULL
                ORDER BY s.start_date DESC, s.start_hour DESC, s.start_minute DESC, s.start_second DESC
            '''
            cursor.execute(query)
            all_data = cursor.fetchall()

            # Add data to Excel
            for row_idx, data in enumerate(all_data, start=2):  # Start at row 2 for headers
                ws[f"A{row_idx}"] = data[0]  # Session ID
                ws[f"B{row_idx}"] = data[1]  # Username
                ws[f"C{row_idx}"] = data[2]  # Start DateTime
                ws[f"D{row_idx}"] = data[3]  # End DateTime
                ws[f"E{row_idx}"] = data[4]  # Time Connected

            try:
                wb.save(self.excel_file)
                logger.info(f"All data added to your Excel file: {self.excel_file}")
            except PermissionError:
                logger.error(f"Please close '{self.excel_file}' and restart the program.")

        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")
        finally:
            if conn:
                conn.close()
